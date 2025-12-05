#!/usr/bin/env python3
"""
Predict NBA spread outcomes for upcoming games
Exports results to Excel with formatted output
"""
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import requests
import sys
import os
import time
import warnings
warnings.filterwarnings('ignore')

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import lightgbm as lgb
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"❌ Missing required library: {e}")
    print("Install with: pip install lightgbm openpyxl")
    sys.exit(1)

NBA_BASE_URL = "http://site.api.espn.com/apis/site/v2/sports/basketball/nba"

# Import pipeline functions
sys.path.insert(0, os.path.dirname(__file__))
from run_full_pipeline import (
    ELORating, calculate_rest_days, calculate_rolling_features, build_features,
    prepare_model_data, load_game_data
)

def fetch_upcoming_games(days_ahead=2):
    """Fetch upcoming games from ESPN API"""
    print(f"\n📅 Fetching upcoming games (next {days_ahead} days)...")
    
    games = []
    today = datetime.now().date()
    
    for i in range(days_ahead):
        date = today + timedelta(days=i)
        date_str = date.strftime('%Y%m%d')
        
        url = f"{NBA_BASE_URL}/scoreboard"
        params = {'dates': date_str}
        
        try:
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            
            if 'events' in data:
                for event in data['events']:
                    if event.get('status', {}).get('type', {}).get('completed', False):
                        continue  # Skip completed games
                    
                    # Parse date from API (comes in UTC)
                    event_date_utc = pd.to_datetime(event.get('date'))
                    # Convert UTC to local time (EST/EDT)
                    if event_date_utc.tzinfo is not None:
                        event_date_local = event_date_utc.tz_convert('America/New_York')
                    else:
                        # If no timezone info, assume UTC and convert
                        event_date_local = event_date_utc.tz_localize('UTC').tz_convert('America/New_York')
                    
                    game_data = {
                        'game_id': event.get('id'),
                        'date': event_date_local.tz_localize(None),  # Convert to local time, then remove timezone
                        'name': event.get('name', ''),
                        'status': event.get('status', {}).get('type', {}).get('name', 'Scheduled'),
                    }
                    
                    # Extract team information
                    for competitor in event.get('competitions', [{}])[0].get('competitors', []):
                        if competitor.get('homeAway') == 'home':
                            game_data.update({
                                'home_team_id': competitor.get('team', {}).get('id'),
                                'home_team_name': competitor.get('team', {}).get('displayName', ''),
                                'home_team_abbrev': competitor.get('team', {}).get('abbreviation', ''),
                                'home_score': competitor.get('score') if competitor.get('score') else None,
                            })
                        else:
                            game_data.update({
                                'away_team_id': competitor.get('team', {}).get('id'),
                                'away_team_name': competitor.get('team', {}).get('displayName', ''),
                                'away_team_abbrev': competitor.get('team', {}).get('abbreviation', ''),
                                'away_score': competitor.get('score') if competitor.get('score') else None,
                            })
                    
                    # Get spread/odds if available (DraftKings data)
                    odds_list = event.get('competitions', [{}])[0].get('odds', [])
                    if odds_list:
                        spread_info = odds_list[0]
                        
                        # Extract opening and closing spreads from pointSpread object
                        point_spread = spread_info.get('pointSpread', {})
                        if point_spread:
                            # Opening spread
                            home_open = point_spread.get('home', {}).get('open', {})
                            if home_open and home_open.get('line'):
                                try:
                                    opening_val = float(home_open.get('line').replace('-', '').replace('+', ''))
                                    if home_open.get('line').startswith('-'):
                                        opening_val = -opening_val
                                    game_data['opening_spread'] = opening_val
                                except:
                                    game_data['opening_spread'] = spread_info.get('spread')
                            else:
                                game_data['opening_spread'] = spread_info.get('spread')
                            
                            # Closing/Current spread (for upcoming games, this is the current line)
                            home_close = point_spread.get('home', {}).get('close', {})
                            if home_close and home_close.get('line'):
                                try:
                                    closing_val = float(home_close.get('line').replace('-', '').replace('+', ''))
                                    if home_close.get('line').startswith('-'):
                                        closing_val = -closing_val
                                    game_data['current_spread'] = closing_val
                                    game_data['closing_spread'] = closing_val  # For upcoming games, current = closing
                                except:
                                    game_data['current_spread'] = spread_info.get('spread')
                                    game_data['closing_spread'] = spread_info.get('spread')
                            else:
                                # Use current spread field as fallback
                                game_data['current_spread'] = spread_info.get('spread')
                                game_data['closing_spread'] = spread_info.get('spread')
                        else:
                            # Fallback to simple spread field
                            game_data['opening_spread'] = spread_info.get('spread')
                            game_data['current_spread'] = spread_info.get('spread')
                            game_data['closing_spread'] = spread_info.get('spread')
                        
                        # Also capture other odds info if available
                        game_data['over_under'] = spread_info.get('overUnder')
                        favorite = spread_info.get('favorite', {})
                        if isinstance(favorite, dict):
                            game_data['favorite_id'] = favorite.get('teamId')
                        else:
                            game_data['favorite_id'] = favorite
                        
                        # Convert to home-team perspective if favorite is away team
                        if game_data.get('favorite_id') and str(game_data['favorite_id']) == str(game_data.get('away_team_id')):
                            if game_data.get('opening_spread') is not None:
                                game_data['opening_spread'] = -game_data['opening_spread']
                            if game_data.get('current_spread') is not None:
                                game_data['current_spread'] = -game_data['current_spread']
                            if game_data.get('closing_spread') is not None:
                                game_data['closing_spread'] = -game_data['closing_spread']
                    
                    games.append(game_data)
            
            time.sleep(0.3)  # Rate limiting
            
        except Exception as e:
            print(f"  ⚠️  Error fetching games for {date}: {e}")
            continue
    
    if not games:
        print("❌ No upcoming games found")
        return None
    
    df = pd.DataFrame(games)
    df['date'] = pd.to_datetime(df['date'])
    df = df.sort_values('date').reset_index(drop=True)
    
    print(f"✓ Found {len(df)} upcoming games")
    return df

def train_model_for_predictions():
    """Train model on historical data"""
    print("\n🔨 Training model on historical data...")
    
    # Load historical data
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(root_dir, 'data')
    
    # Try new naming first, then fallback to old naming
    for filename in [
        os.path.join(data_dir, 'final_dataset_with_injuries.csv'),
        os.path.join(data_dir, 'final_dataset_raw_games.csv'),
        os.path.join(data_dir, 'nba_games_with_injuries.csv'),
        os.path.join(data_dir, 'nba_games.csv')
    ]:
        if os.path.exists(filename):
            data_file = filename
            break
    else:
        print(f"❌ Data file not found in {data_dir}")
        print("   Looking for: final_dataset_with_injuries.csv or nba_games.csv")
        return None, None, None
    
    games_df = pd.read_csv(data_file, parse_dates=['date'])
    games_df = games_df[(games_df['home_score'].notna()) & (games_df['away_score'].notna())].copy()
    
    # Calculate margin before building features
    games_df['margin'] = games_df['home_score'] - games_df['away_score']
    games_df = games_df.sort_values('date').reset_index(drop=True)
    
    # Build features
    features_df = build_features(games_df, include_spread=True)
    
    # Prepare model data
    X, y, feature_names, metadata = prepare_model_data(features_df, target='residual')
    
    # Train model
    from sklearn.model_selection import TimeSeriesSplit
    tscv = TimeSeriesSplit(n_splits=3)
    
    models = []
    best_iters = []
    
    for fold, (train_idx, val_idx) in enumerate(tscv.split(X)):
        X_train, X_val = X[train_idx], X[val_idx]
        y_train, y_val = y[train_idx], y[val_idx]
        
        dtrain = lgb.Dataset(X_train, label=y_train, feature_name=feature_names)
        dval = lgb.Dataset(X_val, label=y_val, reference=dtrain, feature_name=feature_names)
        
        params = {
            "objective": "regression", "metric": "mae", "boosting_type": "gbdt",
            "learning_rate": 0.05, "num_leaves": 31, "feature_fraction": 0.8,
            "bagging_fraction": 0.8, "bagging_freq": 5, "min_child_samples": 20,
            "verbose": -1, "seed": 42
        }
        
        model = lgb.train(params, dtrain, num_boost_round=500,
                         valid_sets=[dval], callbacks=[lgb.early_stopping(stopping_rounds=30, verbose=False)])
        
        models.append(model)
        best_iters.append(model.best_iteration)
    
    # Final model on all data
    avg_best_iter = int(np.mean(best_iters))
    final_model = lgb.train(params, lgb.Dataset(X, label=y, feature_name=feature_names),
                            num_boost_round=avg_best_iter)
    
    # Get ELO ratings at latest date
    elo = ELORating()
    for idx, row in games_df.iterrows():
        home_id = str(row['home_team_id'])
        away_id = str(row['away_team_id'])
        margin = row['home_score'] - row['away_score']
        elo.update(home_id, away_id, margin, row['date'], home_advantage=70)
    
    print(f"✓ Model trained on {len(features_df)} historical games")
    print(f"✓ Using {len(feature_names)} features")
    
    return final_model, feature_names, elo

def build_features_for_upcoming(upcoming_df, historical_df, elo):
    """Build features for upcoming games using historical data"""
    print("\n⚙️  Building features for upcoming games...")
    
    # Combine historical and upcoming (for rolling stats calculation)
    upcoming_df['home_score'] = None
    upcoming_df['away_score'] = None
    upcoming_df['margin'] = None
    
    # Get latest ELO ratings
    latest_date = historical_df['date'].max()
    
    upcoming_features = []
    for idx, row in upcoming_df.iterrows():
        home_id = str(row['home_team_id'])
        away_id = str(row['away_team_id'])
        game_date = row['date']
        
        # Get ELO ratings
        home_elo = elo.get_rating(home_id, latest_date)
        away_elo = elo.get_rating(away_id, latest_date)
        elo_diff = home_elo - away_elo
        
        # Calculate rest days (from historical data)
        home_last_game = historical_df[historical_df['home_team_id'] == row['home_team_id']]
        home_last_game = home_last_game[home_last_game['date'] < game_date]
        if len(home_last_game) > 0:
            last_date = home_last_game['date'].max()
            rest_home = (game_date - last_date).days
        else:
            rest_home = 2  # Default
        
        away_last_game = historical_df[historical_df['away_team_id'] == row['away_team_id']]
        away_last_game = away_last_game[away_last_game['date'] < game_date]
        if len(away_last_game) > 0:
            last_date = away_last_game['date'].max()
            rest_away = (game_date - last_date).days
        else:
            rest_away = 2
        
        rest_diff = rest_home - rest_away
        
        # Simple rolling stats (last 5 games from historical)
        home_recent = historical_df[
            ((historical_df['home_team_id'] == row['home_team_id']) & (historical_df['home_score'].notna())) |
            ((historical_df['away_team_id'] == row['home_team_id']) & (historical_df['away_score'].notna()))
        ].tail(5)
        
        away_recent = historical_df[
            ((historical_df['home_team_id'] == row['away_team_id']) & (historical_df['home_score'].notna())) |
            ((historical_df['away_team_id'] == row['away_team_id']) & (historical_df['away_score'].notna()))
        ].tail(5)
        
        # Calculate rolling averages
        if len(home_recent) > 0:
            home_margins = []
            for _, g in home_recent.iterrows():
                if g['home_team_id'] == row['home_team_id']:
                    home_margins.append(g['home_score'] - g['away_score'])
                else:
                    home_margins.append(g['away_score'] - g['home_score'])
            home_margin_avg = np.mean(home_margins) if home_margins else 0
        else:
            home_margin_avg = 0
        
        if len(away_recent) > 0:
            away_margins = []
            for _, g in away_recent.iterrows():
                if g['home_team_id'] == row['away_team_id']:
                    away_margins.append(g['home_score'] - g['away_score'])
                else:
                    away_margins.append(g['away_score'] - g['home_score'])
            away_margin_avg = np.mean(away_margins) if away_margins else 0
        else:
            away_margin_avg = 0
        
        # Calculate ELO-based spread estimate
        elo_spread_estimate = elo_diff / 25
        
        # Get DraftKings spread (current spread for upcoming games)
        dk_spread = row.get('current_spread') if pd.notna(row.get('current_spread')) else row.get('closing_spread')
        if pd.isna(dk_spread):
            dk_spread = elo_spread_estimate
        
        # Calculate DraftKings spread features
        dk_spread_diff = dk_spread - elo_spread_estimate
        
        # Line movement (opening to current)
        opening_spread = row.get('opening_spread')
        if pd.notna(opening_spread):
            dk_spread_move = dk_spread - opening_spread
        else:
            dk_spread_move = 0
        
        # Build feature row (matching historical feature order)
        feature_row = {
            'elo_home': home_elo,
            'elo_away': away_elo,
            'elo_diff': elo_diff,
            'rest_days_home': rest_home,
            'rest_days_away': rest_away,
            'rest_diff': rest_diff,
            # home_court removed - constant feature causes overfitting
            'day_of_week': game_date.dayofweek,
            'month': game_date.month,
            'is_weekend': 1 if game_date.dayofweek >= 5 else 0,
            'home_margin_avg_3': home_margin_avg,
            'home_margin_avg_5': home_margin_avg,
            'home_margin_avg_10': home_margin_avg,
            'away_margin_avg_3': away_margin_avg,
            'away_margin_avg_5': away_margin_avg,
            'away_margin_avg_10': away_margin_avg,
            'home_points_for_avg_3': 110,  # Placeholder
            'home_points_for_avg_5': 110,
            'home_points_for_avg_10': 110,
            'home_points_against_avg_3': 108,
            'home_points_against_avg_5': 108,
            'home_points_against_avg_10': 108,
            'away_points_for_avg_3': 110,
            'away_points_for_avg_5': 110,
            'away_points_for_avg_10': 110,
            'away_points_against_avg_3': 108,
            'away_points_against_avg_5': 108,
            'away_points_against_avg_10': 108,
            'elo_diff_x_rest_diff': elo_diff * rest_diff,
            # DraftKings spread features
            'dk_spread': dk_spread,
            'dk_spread_diff': dk_spread_diff,
            'dk_spread_move': dk_spread_move,
            'closing_spread': dk_spread,  # Use DraftKings spread as closing_spread for target calculation
        }
        
        upcoming_features.append(feature_row)
    
    features_df = pd.DataFrame(upcoming_features)
    return features_df

def export_to_excel(predictions_df, output_file=None):
    """Export predictions to formatted Excel file"""
    if output_file is None:
        output_file = 'predictions.xlsx'
    print(f"\n📊 Exporting to Excel: {output_file}")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predictions"
    
    # Headers
    # Note:
    # - "Home Spread" is always from the home team's point of view (baseline/Elo line)
    # - "Line" shows sportsbook-style notation (favorite -, underdog +) for that baseline
    # - "Model Line" shows the same notation using the model's predicted margin
    headers = [
        'Date', 'Time', 'Away Team', 'Home Team',
        'Line', 'Home Spread', 'Model Line',
        'Predicted Margin', 'Predicted Residual',
        'Cover Prediction', 'Confidence', 'Recommendation'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Style definitions
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    fill_cover = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_no_cover = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Data rows
    for row_num, (_, pred) in enumerate(predictions_df.iterrows(), 2):
        ws.cell(row=row_num, column=1).value = pred['date'].strftime('%Y-%m-%d') if pd.notna(pred['date']) else ''
        ws.cell(row=row_num, column=2).value = pred['date'].strftime('%I:%M %p') if pd.notna(pred['date']) else ''
        ws.cell(row=row_num, column=3).value = pred['away_team_name']
        ws.cell(row=row_num, column=4).value = pred['home_team_name']
        ws.cell(row=row_num, column=5).value = pred.get('line', '')
        ws.cell(row=row_num, column=6).value = pred.get('home_spread', pred.get('spread', 'N/A'))
        ws.cell(row=row_num, column=7).value = pred.get('model_line', '')
        ws.cell(row=row_num, column=8).value = round(pred['predicted_margin'], 1)
        ws.cell(row=row_num, column=9).value = round(pred['predicted_residual'], 2)
        ws.cell(row=row_num, column=10).value = pred['cover_prediction']
        ws.cell(row=row_num, column=11).value = pred['confidence']
        ws.cell(row=row_num, column=12).value = pred['recommendation']
        
        # Formatting
        for col in range(1, 13):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col > 2 else "left", vertical="center")
            
            if col == 10:  # Cover prediction column
                # Highlight any clear side (favorite or underdog) in green
                if pred['cover_prediction'] in ['Favorite Covers', 'Underdog Covers', 'Home Covers', 'Away Covers']:
                    cell.fill = fill_cover
            
            # Numeric columns: home spread, predicted margin, predicted residual
            if col in [6, 8, 9]:
                cell.number_format = '0.0'
    
    # Auto-adjust column widths
    for col in range(1, 13):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save
    wb.save(output_file)
    print(f"✓ Excel file saved: {output_file}")


def fetch_completed_game_results(game_ids):
    """
    Fetch completed game results from ESPN API for given game_ids.
    Returns a DataFrame with game_id, home_score, away_score, and date.
    """
    if not game_ids or len(game_ids) == 0:
        return pd.DataFrame()
    
    results = []
    print(f"  Fetching results for {len(game_ids)} games from API...")
    
    for game_id in game_ids:
        try:
            url = f"{NBA_BASE_URL}/summary"
            params = {'event': str(game_id)}
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            
            # Extract game info
            event = data.get('header', {}).get('competitions', [{}])[0]
            competitors = event.get('competitors', [])
            
            if len(competitors) >= 2:
                home_team = next((c for c in competitors if c.get('homeAway') == 'home'), None)
                away_team = next((c for c in competitors if c.get('homeAway') == 'away'), None)
                
                if home_team and away_team:
                    home_score = home_team.get('score')
                    away_score = away_team.get('score')
                    
                    # Only include if game is completed with scores
                    if home_score is not None and away_score is not None:
                        results.append({
                            'game_id': str(game_id),
                            'home_score': float(home_score),
                            'away_score': float(away_score),
                            'date': pd.to_datetime(data.get('header', {}).get('competitions', [{}])[0].get('date', ''))
                        })
            
            time.sleep(0.2)  # Rate limiting
            
        except Exception as e:
            # Silently skip if game not found or not completed yet
            continue
    
    if results:
        return pd.DataFrame(results)
    return pd.DataFrame()

def update_prediction_logs(predictions_df):
    """
    Maintain:
      - A rolling latest prediction snapshot (overwritten each run)
      - A growing history of all predictions
      - A correctness file once actual results are available
    """
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(root_dir, 'data')
    predictions_dir = os.path.join(root_dir, "predictions")
    os.makedirs(predictions_dir, exist_ok=True)

    # ------------------------------------------------------------------
    # 1) Save rolling "latest" snapshot (overwrites every run)
    # ------------------------------------------------------------------
    latest_xlsx = os.path.join(predictions_dir, "predictions_latest.xlsx")
    latest_csv = os.path.join(predictions_dir, "predictions_latest.csv")
    export_to_excel(predictions_df, output_file=latest_xlsx)
    predictions_df.to_csv(latest_csv, index=False)
    print(f"✓ Latest predictions saved to: {latest_xlsx} and {latest_csv}")

    # ------------------------------------------------------------------
    # 2) Append to / refresh prediction history (single growing file)
    # ------------------------------------------------------------------
    history_file = os.path.join(predictions_dir, "predictions_history.csv")
    history_xlsx = os.path.join(predictions_dir, "predictions_history.xlsx")
    if os.path.exists(history_file):
        try:
            history_df = pd.read_csv(history_file, parse_dates=['date'])
        except Exception:
            history_df = pd.DataFrame()
    else:
        history_df = pd.DataFrame()

    combined = pd.concat([history_df, predictions_df], ignore_index=True)
    # If we have game_id, normalize to string and de-duplicate so the most
    # recent prediction for a game wins.
    if 'game_id' in combined.columns:
        combined['game_id'] = combined['game_id'].astype(str)
        combined = combined.drop_duplicates(subset=['game_id'], keep='last')
    
    # Sort by date to ensure chronological order
    if 'date' in combined.columns:
        combined['date'] = pd.to_datetime(combined['date'])
        combined = combined.sort_values('date').reset_index(drop=True)
    
    # Save full history
    combined.to_csv(history_file, index=False)

    # ------------------------------------------------------------------
    # 3) Build / refresh correctness tracking file
    # ------------------------------------------------------------------
    if 'game_id' not in combined.columns:
        print("⚠️  game_id not available in predictions; skipping correctness tracking.")
        return
    
    combined['game_id'] = combined['game_id'].astype(str)
    
    # Load historical data if available
    games_df = pd.DataFrame()
    # Try new naming first, then fallback to old naming
    games_file = None
    for filename in [
        os.path.join(data_dir, "final_dataset_with_injuries.csv"),
        os.path.join(data_dir, "final_dataset_raw_games.csv"),
        os.path.join(data_dir, "nba_games_with_injuries.csv"),
        os.path.join(data_dir, "nba_games.csv")
    ]:
        if os.path.exists(filename):
            games_file = filename
            break
    
    if games_file and os.path.exists(games_file):
        try:
            games_df = pd.read_csv(games_file, parse_dates=['date'])
            games_df = games_df[
                (games_df['home_score'].notna()) &
                (games_df['away_score'].notna())
            ].copy()
            if 'game_id' in games_df.columns:
                games_df['game_id'] = games_df['game_id'].astype(str)
        except Exception as e:
            print(f"⚠️  Could not load historical results file: {e}")
            games_df = pd.DataFrame()
    
    # Find predictions that don't have results yet
    if not games_df.empty and 'game_id' in games_df.columns:
        merged_prelim = combined.merge(
            games_df[['game_id', 'home_score', 'away_score']],
            on='game_id',
            how='left',
            suffixes=('', '_from_file')
        )
        # Check if the merge created the expected column
        if 'home_score_from_file' in merged_prelim.columns:
            missing_results = merged_prelim[
                merged_prelim['home_score_from_file'].isna()
            ]['game_id'].unique().tolist()
        else:
            # If merge didn't create the column, all games are missing results
            missing_results = combined['game_id'].dropna().unique().tolist()
    else:
        missing_results = combined['game_id'].dropna().unique().tolist()
    
    # Fetch missing results from API
    api_results_df = pd.DataFrame()
    if missing_results:
        api_results_df = fetch_completed_game_results(missing_results)
    
    # Combine historical file results with API results
    all_results = pd.DataFrame(columns=['game_id', 'home_score', 'away_score'])
    
    if not games_df.empty and 'game_id' in games_df.columns:
        all_results = games_df[['game_id', 'home_score', 'away_score']].copy()
    
    if not api_results_df.empty and all(col in api_results_df.columns for col in ['game_id', 'home_score', 'away_score']):
        if not all_results.empty:
            all_results = pd.concat([
                all_results,
                api_results_df[['game_id', 'home_score', 'away_score']]
            ], ignore_index=True).drop_duplicates(subset=['game_id'], keep='last')
        else:
            all_results = api_results_df[['game_id', 'home_score', 'away_score']].copy()
    
    if all_results.empty:
        print("⚠️  No completed games found that match predictions yet; correctness file will be empty.")
        merged = pd.DataFrame()
        correctness_df = pd.DataFrame(columns=['game_id', 'Date', 'Away Team', 'Home Team', 'Line', 'Prediction', 
                                                'Final Score', 'Home Score', 'Away Score', 'Actual Margin', 
                                                'Predicted Margin', 'Result', 'Correct (1=Yes)'])
    else:
        # Merge predictions with actual results
        merged = combined.merge(
            all_results,
            on='game_id',
            suffixes=('_pred', '_actual'),
            how='left'
        )

        # Drop rows where we still don't have actual scores
        merged = merged[merged['home_score'].notna() & merged['away_score'].notna()].copy()
        if merged.empty:
            print("⚠️  No completed games found that match predictions yet; correctness file will be empty.")
            correctness_df = pd.DataFrame(columns=['game_id', 'Date', 'Away Team', 'Home Team', 'Line', 'Prediction', 
                                                    'Final Score', 'Home Score', 'Away Score', 'Actual Margin', 
                                                    'Predicted Margin', 'Result', 'Correct (1=Yes)'])
        else:
            # Determine actual winner (home/away) and whether prediction was correct.
            merged['actual_margin'] = merged['home_score'] - merged['away_score']
            merged['actual_winner'] = merged['actual_margin'].apply(
                lambda m: 'Home' if m > 0 else ('Away' if m < 0 else 'Push')
            )
            
            # Determine which side was predicted to cover
            # Handle both old format (Home Covers/Away Covers) and new format (Favorite Covers/Underdog Covers)
            def get_predicted_side(row):
                cover_pred = str(row.get('cover_prediction', ''))
                home_spread = row.get('home_spread', row.get('spread', 0))
                
                if cover_pred == 'Home Covers':
                    return 'Home'
                elif cover_pred == 'Away Covers':
                    return 'Away'
                elif cover_pred == 'Favorite Covers':
                    # If home_spread > 0, home is favorite; if < 0, away is favorite
                    return 'Home' if home_spread > 0 else 'Away'
                elif cover_pred == 'Underdog Covers':
                    # If home_spread > 0, away is underdog; if < 0, home is underdog
                    return 'Away' if home_spread > 0 else 'Home'
                else:
                    return 'Unknown'
            
            merged['predicted_side'] = merged.apply(get_predicted_side, axis=1)
            merged['correct'] = (
                (merged['predicted_side'] == merged['actual_winner']) &
                (merged['actual_winner'] != 'Push')
            ).astype(int)

            # Add final score in readable format
            merged['final_score'] = merged.apply(
                lambda row: f"{int(row['home_score'])}-{int(row['away_score'])}" 
                if pd.notna(row['home_score']) and pd.notna(row['away_score']) else None,
                axis=1
            )
            
            # Add result status
            merged['result'] = merged['correct'].apply(
                lambda x: '✓ Correct' if x == 1 else ('✗ Incorrect' if pd.notna(x) else 'Pending')
            )

            # Select a clean set of columns for the correctness file (no duplicates, readable format)
            correctness_cols = [
                'game_id',
                'date' if 'date' in merged.columns else 'date_pred',
                'away_team_name' if 'away_team_name' in merged.columns else 'away_team_name_pred',
                'home_team_name' if 'home_team_name' in merged.columns else 'home_team_name_pred',
                'line' if 'line' in merged.columns else None,
                'cover_prediction',
                'final_score',
                'home_score',
                'away_score',
                'actual_margin',
                'predicted_margin',
                'result',
                'correct',
            ]
            
            # Filter out None values and columns that don't exist
            correctness_cols = [c for c in correctness_cols if c and c in merged.columns]
            correctness_df = merged[correctness_cols].copy()
    
            # Rename columns for clarity
            correctness_df = correctness_df.rename(columns={
                'date': 'Date',
                'date_pred': 'Date',
                'away_team_name': 'Away Team',
                'away_team_name_pred': 'Away Team',
                'home_team_name': 'Home Team',
                'home_team_name_pred': 'Home Team',
                'line': 'Line',
                'cover_prediction': 'Prediction',
                'final_score': 'Final Score',
                'home_score': 'Home Score',
                'away_score': 'Away Score',
                'actual_margin': 'Actual Margin',
                'predicted_margin': 'Predicted Margin',
                'result': 'Result',
                'correct': 'Correct (1=Yes)',
            })

            # Sort correctness by date chronologically
            if 'Date' in correctness_df.columns:
                correctness_df['Date'] = pd.to_datetime(correctness_df['Date'])
                correctness_df = correctness_df.sort_values('Date').reset_index(drop=True)
    
    correctness_file = os.path.join(predictions_dir, "prediction_correctness.csv")
    correctness_xlsx = os.path.join(predictions_dir, "prediction_correctness.xlsx")
    correctness_df.to_csv(correctness_file, index=False)
    try:
        correctness_df.to_excel(correctness_xlsx, index=False)
        print(f"✓ Prediction correctness updated: {correctness_file} and {correctness_xlsx}")
    except Exception as e:
        print(f"⚠️  Could not write Excel correctness file: {e}")

    # ------------------------------------------------------------------
    # 4) Clean up history and add correctness info for summary
    # ------------------------------------------------------------------
    # Remove duplicate columns from history (keep only one of spread/home_spread)
    history_cols_to_keep = [
        'game_id', 'date', 'away_team_name', 'home_team_name', 
        'away_team_abbrev', 'home_team_abbrev',
        'line', 'model_line', 'home_spread',  # Keep home_spread, drop 'spread' duplicate
        'predicted_margin', 'predicted_residual',
        'cover_prediction', 'cover_team', 'favorite_team', 'underdog_team',
        'confidence', 'recommendation'
    ]
    
    # Only keep columns that exist
    history_cols_to_keep = [c for c in history_cols_to_keep if c in combined.columns]
    history_clean = combined[history_cols_to_keep].copy()
    
    # Merge correctness info into history for summary
    if not merged.empty and 'home_score' in merged.columns:
        # Create lookup from merged dataframe (before renaming)
        correctness_lookup_raw = merged.set_index('game_id')[['home_score', 'away_score', 'correct']].to_dict('index')
        
        history_clean['home_score'] = history_clean['game_id'].map(
            lambda x: correctness_lookup_raw.get(str(x), {}).get('home_score', None)
        )
        history_clean['away_score'] = history_clean['game_id'].map(
            lambda x: correctness_lookup_raw.get(str(x), {}).get('away_score', None)
        )
        history_clean['Final Score'] = history_clean.apply(
            lambda row: f"{int(row['home_score'])}-{int(row['away_score'])}" 
            if pd.notna(row.get('home_score')) and pd.notna(row.get('away_score')) else None,
            axis=1
        )
        history_clean['Result'] = history_clean['game_id'].map(
            lambda x: '✓ Correct' if correctness_lookup_raw.get(str(x), {}).get('correct') == 1 
            else ('✗ Incorrect' if correctness_lookup_raw.get(str(x), {}).get('correct') == 0 else 'Pending')
        )
        history_clean['Correct'] = history_clean['game_id'].map(
            lambda x: correctness_lookup_raw.get(str(x), {}).get('correct', None)
        )
    else:
        history_clean['home_score'] = None
        history_clean['away_score'] = None
        history_clean['Final Score'] = None
        history_clean['Result'] = 'Pending'
        history_clean['Correct'] = None
    
    # Rename columns for clarity
    history_clean = history_clean.rename(columns={
        'date': 'Date',
        'away_team_name': 'Away Team',
        'home_team_name': 'Home Team',
        'away_team_abbrev': 'Away',
        'home_team_abbrev': 'Home',
        'line': 'Line',
        'model_line': 'Model Line',
        'home_spread': 'Home Spread',
        'predicted_margin': 'Predicted Margin',
        'predicted_residual': 'Predicted Residual',
        'cover_prediction': 'Cover Prediction',
        'cover_team': 'Cover Team',
        'favorite_team': 'Favorite',
        'underdog_team': 'Underdog',
        'confidence': 'Confidence',
        'recommendation': 'Recommendation',
    })
    
    # Ensure Date column is sorted chronologically
    if 'Date' in history_clean.columns:
        history_clean['Date'] = pd.to_datetime(history_clean['Date'])
        history_clean = history_clean.sort_values('Date').reset_index(drop=True)
    
    # Save cleaned history Excel file
    try:
        history_clean.to_excel(history_xlsx, index=False, engine='openpyxl')
        print(f"✓ Prediction history updated: {history_file} and {history_xlsx}")
    except Exception as e:
        print(f"⚠️  Could not write Excel history file: {e}")
        import traceback
        traceback.print_exc()

    # ------------------------------------------------------------------
    # 5) Combined summary workbook with both history and correctness
    # ------------------------------------------------------------------
    summary_xlsx = os.path.join(predictions_dir, "predictions_summary.xlsx")
    try:
        with pd.ExcelWriter(summary_xlsx, engine="openpyxl") as writer:
            history_clean.to_excel(writer, sheet_name="History", index=False)
            if not correctness_df.empty:
                correctness_df.to_excel(writer, sheet_name="Correctness", index=False)
            else:
                # Create empty correctness sheet with proper columns
                empty_correctness = pd.DataFrame(columns=['game_id', 'Date', 'Away Team', 'Home Team', 'Line', 
                                                          'Prediction', 'Final Score', 'Home Score', 'Away Score', 
                                                          'Actual Margin', 'Predicted Margin', 'Result', 'Correct (1=Yes)'])
                empty_correctness.to_excel(writer, sheet_name="Correctness", index=False)
        print(f"✓ Prediction summary workbook updated: {summary_xlsx}")
    except Exception as e:
        print(f"⚠️  Could not write summary workbook: {e}")
        import traceback
        traceback.print_exc()

def main():
    print("="*70)
    print("NBA SPREAD PREDICTIONS - UPCOMING GAMES")
    print("="*70)
    
    # Fetch upcoming games
    upcoming_df = fetch_upcoming_games(days_ahead=2)
    if upcoming_df is None or len(upcoming_df) == 0:
        print("❌ No upcoming games to predict")
        return
    
    # Train model
    model, feature_names, elo = train_model_for_predictions()
    if model is None:
        return
    
    # Load historical data for feature building
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(root_dir, 'data')
    
    # Try new naming first, then fallback to old naming
    historical_file = None
    for filename in [
        os.path.join(data_dir, 'final_dataset_with_injuries.csv'),
        os.path.join(data_dir, 'final_dataset_raw_games.csv'),
        os.path.join(data_dir, 'nba_games_with_injuries.csv'),
        os.path.join(data_dir, 'nba_games.csv')
    ]:
        if os.path.exists(filename):
            historical_file = filename
            break
    
    if not historical_file:
        print(f"❌ Historical data file not found in {data_dir}")
        return
    
    historical_df = pd.read_csv(historical_file, parse_dates=['date'])
    
    # Build features for upcoming games
    features_df = build_features_for_upcoming(upcoming_df, historical_df, elo)
    
    # Prepare prediction data
    exclude_cols = ['closing_spread']  # Keep for reference
    feature_cols = [c for c in features_df.columns if c not in exclude_cols]
    
    # Match feature order
    X_pred = features_df[feature_cols].fillna(0)
    
    # Align features
    missing_features = set(feature_names) - set(X_pred.columns)
    extra_features = set(X_pred.columns) - set(feature_names)
    
    for feat in missing_features:
        X_pred[feat] = 0
    
    X_pred = X_pred[feature_names].fillna(0).values
    
    # Make predictions
    print("\n🎯 Making predictions...")
    residuals = model.predict(X_pred)
    
    # Create predictions dataframe
    predictions = []
    for idx, row in upcoming_df.iterrows():
        residual = residuals[idx]
        spread = features_df.iloc[idx]['closing_spread']
        predicted_margin = spread + residual
        # Use a rounded home-spread value for display/notation
        home_spread = round(spread, 1)
        
        # Sportsbook-style line notation
        home_abbrev = row.get('home_team_abbrev', '') or row.get('home_team_name', 'Home')
        away_abbrev = row.get('away_team_abbrev', '') or row.get('away_team_name', 'Away')
        if home_spread > 0:
            # Home is favorite: HOME -X.X, Away +X.X
            line_str = f"{home_abbrev} -{abs(home_spread):.1f} / {away_abbrev} +{abs(home_spread):.1f}"
        elif home_spread < 0:
            # Home is underdog: HOME +X.X, Away -X.X (favorite)
            line_str = f"{home_abbrev} +{abs(home_spread):.1f} / {away_abbrev} -{abs(home_spread):.1f}"
        else:
            # Pick'em
            line_str = f"{home_abbrev} PK / {away_abbrev} PK"

        # Model-implied line using predicted margin (home-centric)
        model_home_spread = round(predicted_margin, 1)
        if model_home_spread > 0:
            model_line_str = f"{home_abbrev} -{abs(model_home_spread):.1f} / {away_abbrev} +{abs(model_home_spread):.1f}"
        elif model_home_spread < 0:
            model_line_str = f"{home_abbrev} +{abs(model_home_spread):.1f} / {away_abbrev} -{abs(model_home_spread):.1f}"
        else:
            model_line_str = f"{home_abbrev} PK / {away_abbrev} PK"
        
        # Cover prediction, expressed in favorite/underdog terms
        # Use optimal threshold (1.40) - found through model testing with more regularization
        # This threshold balances predictions and reduces Home Cover bias
        COVER_THRESHOLD = 1.40
        
        if home_spread == 0:
            # Pick'em: fall back to home/away language
            cover_pred = 'Home Covers' if residual > COVER_THRESHOLD else 'Away Covers'
            cover_side_team = home_abbrev if residual > COVER_THRESHOLD else away_abbrev
        else:
            # Determine favorite vs underdog
            favorite_is_home = home_spread > 0
            home_covers = residual > COVER_THRESHOLD  # from home-team perspective

            if favorite_is_home:
                # Home is favorite
                if home_covers:
                    cover_pred = 'Favorite Covers'
                    cover_side_team = home_abbrev
                else:
                    cover_pred = 'Underdog Covers'
                    cover_side_team = away_abbrev
            else:
                # Away is favorite
                if home_covers:
                    cover_pred = 'Underdog Covers'
                    cover_side_team = home_abbrev
                else:
                    cover_pred = 'Favorite Covers'
                    cover_side_team = away_abbrev
        
        # Confidence (based on residual magnitude)
        abs_residual = abs(residual)
        if abs_residual > 3:
            confidence = 'High'
        elif abs_residual > 1.5:
            confidence = 'Medium'
        else:
            confidence = 'Low'
        
        # Recommendation
        if abs_residual > 3:
            recommendation = f"{cover_pred} (Strong)"
        elif abs_residual > 1.5:
            recommendation = f"{cover_pred} (Moderate)"
        else:
            recommendation = 'No strong edge'
        
        predictions.append({
            'game_id': row.get('game_id', None),
            'home_team_id': row.get('home_team_id', None),
            'away_team_id': row.get('away_team_id', None),
            'date': row['date'],
            'away_team_name': row['away_team_name'],
            'home_team_name': row['home_team_name'],
            'away_team_abbrev': row.get('away_team_abbrev', ''),
            'home_team_abbrev': row.get('home_team_abbrev', ''),
            # Always from home team's perspective
            'home_spread': home_spread,
            'spread': home_spread,  # keep legacy column name for compatibility
            'line': line_str,
            'model_home_spread': model_home_spread,
            'model_line': model_line_str,
            'favorite_team': home_abbrev if home_spread > 0 else (away_abbrev if home_spread < 0 else 'PK'),
            'underdog_team': away_abbrev if home_spread > 0 else (home_abbrev if home_spread < 0 else 'PK'),
            'cover_team': cover_side_team,
            'predicted_margin': predicted_margin,
            'predicted_residual': residual,
            'cover_prediction': cover_pred,
            'confidence': confidence,
            'recommendation': recommendation,
        })
    
    predictions_df = pd.DataFrame(predictions)

    # Display results
    print("\n" + "="*70)
    print("PREDICTIONS")
    print("="*70)
    print(f"\n{'Date':<12} {'Matchup':<30} {'Line':<22} {'Model Line':<22} {'Pred Margin':<12} {'Cover':<15} {'Confidence':<10}")
    print("-"*70)
    
    for _, pred in predictions_df.iterrows():
        matchup = f"{pred['away_team_abbrev']} @ {pred['home_team_abbrev']}"
        date_str = pred['date'].strftime('%m/%d %I:%M%p')
        line_str = pred.get('line', '')
        model_line_str = pred.get('model_line', '')
        print(f"{date_str:<12} {matchup:<30} {line_str:<22} {model_line_str:<22} {pred['predicted_margin']:>9.1f}  "
              f"{pred['cover_prediction']:<15} {pred['confidence']:<10}")
    
    # Save latest / history / correctness files
    update_prediction_logs(predictions_df)
    
    print("\n" + "="*70)
    print("✓ Predictions complete!")
    print("="*70)

if __name__ == '__main__':
    import time
    main()

